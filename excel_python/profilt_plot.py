import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt


def pro_plot(filename):
	wb = xl.load_workbook(filename)
	sheet = wb['Sheet1']
	# for row in range(2,sheet.max_row+1):
	# 	cell = sheet.cell(row,3)
	# 	# print(cell.value)

		# corrected_price_cell = sheet.cell(row,4)
		# corrected_price_cell.value = cell.value*0.9

	values = Reference(sheet,
				min_row = 2,
				max_row = sheet.max_row,
				min_col = 3,
				max_col = 3)

	chart = BarChart()
	chart.add_data(values)
	sheet.add_chart(chart,"E2")

	wb.save(filename)


filename = 'share_transactions.xlsx'

pro_plot(filename)